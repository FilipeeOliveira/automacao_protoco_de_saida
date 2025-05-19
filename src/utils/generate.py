import os
import sys
from datetime import datetime
from openpyxl import load_workbook

def get_paths():
    """Define os caminhos corretos para ambos os ambientes"""
    if getattr(sys, 'frozen', False):
        # Modo executável
        base_dir = os.path.dirname(sys.executable)
        return {
            'cache_dir': os.path.join(base_dir, "cache_control"),
            'template_dir': os.path.join(base_dir, "template_base")
        }
    else:
        # Modo desenvolvimento
        script_dir = os.path.dirname(os.path.abspath(__file__))
        
        # Verifica se já estamos em src/ ou subpastas
        if os.path.basename(os.path.dirname(script_dir)) == 'src':
            base_dir = os.path.dirname(script_dir)
        else:
            base_dir = script_dir
            
        return {
            'cache_dir': os.path.join(base_dir, "cache_control"),
            'template_dir': os.path.join(base_dir, "template_base")
        }

# Obtém os caminhos
paths = get_paths()
LAST_CONTROL_FILE = os.path.join(paths['cache_dir'], "last_control.txt")
TEMPLATE_PATH = os.path.join(paths['template_dir'], "TERMO_DE_ENTREGA_DE_EQUIPAMENTO.xlsx")

# Cria diretórios se não existirem
os.makedirs(paths['cache_dir'], exist_ok=True)
os.makedirs(paths['template_dir'], exist_ok=True)

# Verifica se o template existe
if not os.path.exists(TEMPLATE_PATH):
    available = "\n  • " + "\n  • ".join(os.listdir(paths['template_dir'])) if os.path.exists(paths['template_dir']) else "Diretório não existe"
    raise FileNotFoundError(
        f"Template não encontrado em: {TEMPLATE_PATH}\n"
        f"Arquivos disponíveis no diretório:{available}"
    )

def load_last_control():
    if os.path.exists(LAST_CONTROL_FILE):
        return open(LAST_CONTROL_FILE).read().strip()
    return None

def save_last_control(control):
    with open(LAST_CONTROL_FILE, 'w') as f:
        f.write(control)

def get_next_control():
    last = load_last_control()
    year = datetime.now().year
    if last and '/' in last:
        num_str, y_str = last.split('/')
        try:
            num, y = int(num_str), int(y_str)
        except ValueError:
            num, y = 0, year
        next_num = num + 1 if y == year else 1
    else:
        next_num = 1
    return f"{next_num:04d}/{year}"

def fill_term(data: dict, equipamentos: list, output_dir: str):
    """
    data: dict com chaves correspondentes a todas as células (controle, nome, cpf, etc.)
    equipamentos: lista de dicts com descricao, quantidade, unidade, estoque
    """
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["ORDEM DE RETIRADA DE ESTOQUE"]

    # Mapeia cada campo do data para a célula certa:
    mapping = {
        'controle': 'G2',
        'local_saida': 'C3',
        'local_destino': 'C4',
        'data_saida': 'F4',
        'motivo': 'C5',
        'data_retorno': 'F5',
        'nome': 'B7',
        'cpf': 'E7',
        'setor': 'B8',
        'cargo': 'E8',
        'responsavel_setor': 'B9',
        'observacao': 'A23',
    }
    for key, cell in mapping.items():
        ws[cell] = data.get(key, "")

    # Insere equipamentos a partir da linha 12
    row = 12
    for eq in equipamentos:
        ws[f"A{row}"] = eq['descricao']
        ws[f"E{row}"] = eq['quantidade']
        ws[f"F{row}"] = eq['unidade']
        ws[f"G{row}"] = eq['estoque']
        row += 1

    # Data de impressão nas células fixas:
    hoje = datetime.now().strftime("%d-%m-%Y")
    for c in ['A32','C32','D32','F32']:
        ws[c] = f"DATA: {hoje}"

    # Cria pasta de saída se não existir
    os.makedirs(output_dir, exist_ok=True)
    filename = f"Termo_{data['nome']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    fullpath = os.path.join(output_dir, filename)
    wb.save(fullpath)

    # atualiza controle
    save_last_control(data['controle'])
    return fullpath
