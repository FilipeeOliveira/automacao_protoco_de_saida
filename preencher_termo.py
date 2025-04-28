import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry  
from openpyxl import load_workbook
from datetime import datetime
import os, sys, subprocess

class TermoEntregaApp:
    def __init__(self, root, output_dir="termos_salvos"):
        self.root = root
        self.root.title("Preenchimento de Termo de Entrega")
        # Diretório onde os arquivos serão salvos
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Variáveis para armazenar os dados
        self.controle_var = tk.StringVar(value=self.get_next_control())
        self.local_saida_var = tk.StringVar(value="ANTONELLY SEDE - ESTOQUE T.I")
        self.local_destino_var = tk.StringVar()
        self.motivo_var = tk.StringVar()
        self.data_saida_var = tk.StringVar()
        self.data_retorno_var = tk.StringVar()
        self.nome_var = tk.StringVar()
        self.cpf_var = tk.StringVar()
        self.setor_var = tk.StringVar()
        self.cargo_var = tk.StringVar()
        self.responsavel_setor_var = tk.StringVar()
        self.descricao_var = tk.StringVar()
        self.quantidade_var = tk.IntVar(value=1)
        self.unidade_var = tk.StringVar()
        self.estoque_var = tk.StringVar(value="SEDE")
        self.observacao_var = tk.StringVar()
        
        # Lista para armazenar múltiplos equipamentos
        self.equipamentos = []
        
        # Criar a interface
        self.create_widgets()

    def load_last_control(self):
        path = 'last_control.txt'
        if os.path.exists(path):
            with open(path, 'r') as f:
                return f.read().strip()
        return None

    def save_last_control(self, control):
        with open('last_control.txt', 'w') as f:
            f.write(control)

    def get_next_control(self):
        last = self.load_last_control()
        year = datetime.now().year
        if last and '/' in last:
            num_str, y_str = last.split('/')
            try:
                num = int(num_str)
                y = int(y_str)
            except ValueError:
                num, y = 0, year
            next_num = num + 1 if y == year else 1
        else:
            next_num = 1
        return f"{next_num:04d}/{year}"

    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.columnconfigure((0,1), weight=1)
        
        # Informações do Termo
        header_frame = ttk.LabelFrame(main_frame, text="Informações do Termo", padding=10)
        header_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        header_frame.columnconfigure(1, weight=1)

        # Nº Controle
        ttk.Label(header_frame, text="Nº de Controle:").grid(row=0, column=0, sticky="e", pady=2)
        ttk.Entry(header_frame, textvariable=self.controle_var).grid(row=0, column=1, sticky="ew", pady=2)
        # Local de Saída
        ttk.Label(header_frame, text="Local de Saída:").grid(row=1, column=0, sticky="e", pady=2)
        ttk.Entry(header_frame, textvariable=self.local_saida_var).grid(row=1, column=1, sticky="ew", pady=2)
        # Local de Destino como Combobox
        ttk.Label(header_frame, text="Local de Destino:").grid(row=2, column=0, sticky="e", pady=2)
        destino_combo = ttk.Combobox(header_frame, textvariable=self.local_destino_var)
        destino_combo['values'] = ["Macapa", "DNIT", "HOME - OFFICE"]
        destino_combo.state(['!readonly'])
        destino_combo.grid(row=2, column=1, sticky="ew", pady=2)
        # Motivo
        ttk.Label(header_frame, text="Motivo:").grid(row=3, column=0, sticky="e", pady=2)
        ttk.Entry(header_frame, textvariable=self.motivo_var).grid(row=3, column=1, sticky="ew", pady=2)
        # Datas
        ttk.Label(header_frame, text="Data de Saída:").grid(row=4, column=0, sticky="e", pady=2)
        DateEntry(header_frame, textvariable=self.data_saida_var, date_pattern='dd-MM-yyyy').grid(row=4, column=1, sticky="w", pady=2)
        ttk.Label(header_frame, text="Data de Retorno:").grid(row=5, column=0, sticky="e", pady=2)
        DateEntry(header_frame, textvariable=self.data_retorno_var, date_pattern='dd-MM-yyyy').grid(row=5, column=1, sticky="w", pady=2)

        # Responsável com Combobox para Setor e Cargo
        responsavel_frame = ttk.LabelFrame(main_frame, text="Responsável", padding=10)
        responsavel_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
        responsavel_frame.columnconfigure(1, weight=1)
        labels_r = ["Nome:", "CPF:", "Setor:", "Cargo:", "Responsável do Setor:"]
        for i, lbl in enumerate(labels_r):
            ttk.Label(responsavel_frame, text=lbl).grid(row=i, column=0, sticky="e", pady=2)
        ttk.Entry(responsavel_frame, textvariable=self.nome_var).grid(row=0, column=1, sticky="ew", pady=2)
        ttk.Entry(responsavel_frame, textvariable=self.cpf_var).grid(row=1, column=1, sticky="ew", pady=2)
        setor_combo = ttk.Combobox(responsavel_frame, textvariable=self.setor_var)
        setor_combo['values'] = [
            "Engenharia", "Financeiro", "Pagamentos",
            "Compras", "Jurídico", "RH", "Marketing"
        ]
        setor_combo.state(['!readonly'])
        setor_combo.grid(row=2, column=1, sticky="ew", pady=2)
        cargo_combo = ttk.Combobox(responsavel_frame, textvariable=self.cargo_var)
        cargo_combo['values'] = [
            "Assistente Administrativo", "Analista de TI", "Coordenador de Projetos",
            "Diretor Financeiro", "Gerente de Compras", "Advogado", "Auxiliar de RH", "Analista de Marketing"
        ]
        cargo_combo.state(['!readonly'])
        cargo_combo.grid(row=3, column=1, sticky="ew", pady=2)
        ttk.Entry(responsavel_frame, textvariable=self.responsavel_setor_var).grid(row=4, column=1, sticky="ew", pady=2)

        # Equipamentos
        equipamento_frame = ttk.LabelFrame(main_frame, text="Detalhamento do Equipamento", padding=10)
        equipamento_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)
        equipamento_frame.columnconfigure(1, weight=1)
        eq_labels = ["Descrição:", "Quantidade:", "Unidade:", "Estoque:"]
        eq_vars = [self.descricao_var, self.quantidade_var, self.unidade_var, self.estoque_var]
        for i, (lbl, var) in enumerate(zip(eq_labels, eq_vars)):
            ttk.Label(equipamento_frame, text=lbl).grid(row=i, column=0, sticky="e", pady=2)
            if lbl == "Quantidade:":
                ttk.Spinbox(equipamento_frame, from_=1, to=9999, textvariable=self.quantidade_var).grid(row=i, column=1, sticky="w", pady=2)
            else:
                ttk.Entry(equipamento_frame, textvariable=var, width=50 if lbl=="Descrição:" else None).grid(row=i, column=1, sticky="ew", pady=2)
        btn_frame = ttk.Frame(equipamento_frame)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=5)
        ttk.Button(btn_frame, text="Adicionar Equipamento", command=self.adicionar_equipamento).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Remover Selecionado", command=self.remover_equipamento).pack(side=tk.LEFT, padx=5)
        self.lista_equipamentos = tk.Listbox(equipamento_frame, height=4)
        self.lista_equipamentos.grid(row=5, column=0, columnspan=2, sticky="ew", pady=5)
        ttk.Label(equipamento_frame, text="Observação:").grid(row=6, column=0, sticky="e", pady=2)
        ttk.Entry(equipamento_frame, textvariable=self.observacao_var, width=50).grid(row=6, column=1, sticky="ew", pady=2)
        
        # Ações
        action_frame = ttk.Frame(main_frame)
        action_frame.grid(row=2, column=0, columnspan=2, pady=10)
        ttk.Button(action_frame, text="Preencher Termo", command=self.preencher_termo).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="Limpar", command=self.limpar_campos).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="Termos Criados", command=self.open_folder).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="Sair", command=self.root.quit).pack(side=tk.LEFT, padx=5)

    def open_folder(self):
        """Abre a pasta de destino no explorador de arquivos"""
        path = os.path.abspath(self.output_dir)
        try:
            if sys.platform.startswith('darwin'):
                subprocess.call(['open', path])
            elif sys.platform.startswith('linux'):
                subprocess.call(['xdg-open', path])
            else:
                os.startfile(path)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível abrir a pasta:\n{e}")

    def adicionar_equipamento(self):
        descricao = self.descricao_var.get()
        quantidade = self.quantidade_var.get()
        unidade = self.unidade_var.get()
        estoque = self.estoque_var.get()
        if not descricao:
            messagebox.showwarning("Aviso", "Digite uma descrição para o equipamento!")
            return
        equipamento = {"descricao": descricao, "quantidade": quantidade, "unidade": unidade, "estoque": estoque}
        self.equipamentos.append(equipamento)
        self.lista_equipamentos.insert(tk.END, f"{quantidade} {unidade} - {descricao} ({estoque})")
        self.descricao_var.set("")
        self.quantidade_var.set(1)
        self.unidade_var.set("UND")
        self.estoque_var.set("SEDE")

    def remover_equipamento(self):
        selecionado = self.lista_equipamentos.curselection()
        if selecionado:
            self.equipamentos.pop(selecionado[0])
            self.lista_equipamentos.delete(selecionado[0])

    def preencher_termo(self):
        if not self.equipamentos:
            messagebox.showwarning("Aviso", "Adicione pelo menos um equipamento!")
            return
        try:
            template_path = "TERMO_DE_ENTREGA_DE_EQUIPAMENTO_151.xlsx"
            wb = load_workbook(template_path)
            ws = wb["ORDEM DE RETIRADA DE ESTOQUE"]
            ws['G2'] = self.controle_var.get()
            ws['C3'] = self.local_saida_var.get()
            ws['C4'] = self.local_destino_var.get()
            ws['F4'] = self.data_saida_var.get()
            ws['C5'] = self.motivo_var.get()
            ws['F5'] = self.data_retorno_var.get()
            ws['B7'] = self.nome_var.get()
            ws['E7'] = self.cpf_var.get()
            ws['B8'] = self.setor_var.get()
            ws['E8'] = self.cargo_var.get()
            ws['B9'] = self.responsavel_setor_var.get()
            ws['A23'] = self.observacao_var.get()
            linha = 12
            for eq in self.equipamentos:
                ws[f'A{linha}'] = eq['descricao']
                ws[f'E{linha}'] = eq['quantidade']
                ws[f'F{linha}'] = eq['unidade']
                ws[f'G{linha}'] = eq['estoque']
                linha += 1
            data_atual = datetime.now().strftime("%d-%m-%Y")
            for cell in ['A32','C32','D32','F32']:
                ws[cell] = f"DATA: {data_atual}"
            full_path = os.path.join(self.output_dir, f"Termo_de_Saída{self.nome_var.get()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            wb.save(full_path)
            messagebox.showinfo("Sucesso", f"Termo salvo em:\n{full_path}")
            self.save_last_control(self.controle_var.get())
            self.limpar_campos()
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao preencher o termo:\n{e}")

    def limpar_campos(self):
        self.controle_var.set(self.get_next_control())
        self.nome_var.set("")
        self.cpf_var.set("")
        self.setor_var.set("")
        self.cargo_var.set("")
        self.responsavel_setor_var.set("")
        self.descricao_var.set("")
        self.quantidade_var.set(1)
        self.unidade_var.set("UND")
        self.estoque_var.set("SEDE")
        self.observacao_var.set("")
        self.data_saida_var.set("")
        self.data_retorno_var.set("")
        self.equipamentos = []
        self.lista_equipamentos.delete(0, tk.END)

if __name__ == "__main__":
    root = tk.Tk()
    app = TermoEntregaApp(root)
    root.mainloop()